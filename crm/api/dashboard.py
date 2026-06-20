import json

import frappe
from frappe import _

from crm.fcrm.doctype.crm_dashboard.crm_dashboard import create_default_manager_dashboard
from crm.utils import sales_user_only


def get_department_users(department):
	"""Get all users in a department from Department Pipeline Stage team members."""
	if not department:
		return []

	try:
		doc = frappe.get_doc("Department Pipeline Stage", department)
		return [tm.user for tm in (doc.team_members or []) if tm.user]
	except frappe.DoesNotExistError:
		return []


@frappe.whitelist()
def get_departments():
	"""Get all department pipeline stages for the frontend dropdown."""
	depts = frappe.get_all(
		"Department Pipeline Stage",
		fields=["name"],
		order_by="name",
	)
	return depts


@frappe.whitelist()
def get_department_users_list(department):
	"""Get users for a department - used by frontend to show user pills."""
	users = get_department_users(department)
	result = []
	for u in users:
		full_name = frappe.db.get_value("User", u, "full_name") or u
		result.append({"name": u, "full_name": full_name})
	return result


@frappe.whitelist()
def reset_to_default():
	frappe.only_for("System Manager")
	create_default_manager_dashboard(force=True)


@frappe.whitelist()
def force_reset_layout():
	from crm.fcrm.doctype.crm_dashboard.crm_dashboard import default_manager_dashboard_layout
	layout = default_manager_dashboard_layout()
	frappe.db.set_value("CRM Dashboard", "Manager Dashboard", "layout", layout)
	frappe.db.commit()
	return "Success"


@frappe.whitelist()
@sales_user_only
def get_dashboard(from_date="", to_date="", user="", department=""):
	"""
	Get the dashboard data for the CRM dashboard.
	"""

	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	roles = frappe.get_roles(frappe.session.user)
	is_sales_manager = "Sales Manager" in roles or "System Manager" in roles
	is_sales_user = "Sales User" in roles and not is_sales_manager

	# Build users list from department or single user
	users = []
	if is_sales_user:
		users = [frappe.session.user]
	elif department and not user:
		users = get_department_users(department)
		# If department selected but has no users, use sentinel to return zero results
		if not users:
			users = ["__no_match__"]
	elif user:
		users = [user]

	dashboard = frappe.db.exists("CRM Dashboard", "Manager Dashboard")

	layout = []

	if not dashboard:
		layout = json.loads(create_default_manager_dashboard())
		frappe.db.commit()
	else:
		layout = json.loads(frappe.db.get_value("CRM Dashboard", "Manager Dashboard", "layout") or "[]")

	for l in layout:
		method_name = f"get_{l['name']}"
		try:
			if hasattr(frappe.get_attr("crm.api.dashboard"), method_name):
				method = getattr(frappe.get_attr("crm.api.dashboard"), method_name)
				l["data"] = method(from_date, to_date, users)
			else:
				l["data"] = None
		except Exception as e:
			frappe.log_error(f"Dashboard component {l['name']} failed: {str(e)}", "CRM Dashboard Error")
			l["data"] = None

	return layout


@frappe.whitelist()
@sales_user_only
def get_chart(name, type, from_date="", to_date="", user="", department=""):
	"""
	Get number chart data for the dashboard.
	"""
	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	roles = frappe.get_roles(frappe.session.user)
	is_sales_manager = "Sales Manager" in roles or "System Manager" in roles
	is_sales_user = "Sales User" in roles and not is_sales_manager

	users = []
	if is_sales_user:
		users = [frappe.session.user]
	elif department and not user:
		users = get_department_users(department)
		if not users:
			users = ["__no_match__"]
	elif user:
		users = [user]

	method_name = f"get_{name}"
	if hasattr(frappe.get_attr("crm.api.dashboard"), method_name):
		method = getattr(frappe.get_attr("crm.api.dashboard"), method_name)
		return method(from_date, to_date, users)
	else:
		return {"error": _("Invalid chart name")}


def get_total_leads(from_date, to_date, users=None):
	"""
	Get lead count for the dashboard.
	"""
	conds = ""

	diff = frappe.utils.date_diff(to_date, from_date)
	if diff == 0:
		diff = 1
	params = {
		"from_date": from_date,
		"to_date": to_date,
		"prev_from_date": frappe.utils.add_days(from_date, -diff),
	}

	if users:
		conds += " AND lead_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
            COUNT(CASE
                WHEN creation >= %(from_date)s AND creation < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
                {conds}
                THEN name
                ELSE NULL
            END) as current_month_leads,

            COUNT(CASE
                WHEN creation >= %(prev_from_date)s AND creation < %(from_date)s
                {conds}
                THEN name
                ELSE NULL
            END) as prev_month_leads
		FROM `tabCRM Lead`
	""",
		params,
		as_dict=1,
	)

	current_month_leads = result[0].current_month_leads or 0
	prev_month_leads = result[0].prev_month_leads or 0

	delta_in_percentage = (
		(current_month_leads - prev_month_leads) / prev_month_leads * 100 if prev_month_leads else 0
	)

	return {
		"title": _("Total leads"),
		"tooltip": _("Total number of leads"),
		"value": current_month_leads,
		"delta": delta_in_percentage,
		"deltaSuffix": "%",
	}


def get_ongoing_deals(from_date, to_date, users=None):
	"""
	Get ongoing deal count for the dashboard, and also calculate average deal value for ongoing deals.
	"""
	conds = ""

	diff = frappe.utils.date_diff(to_date, from_date)
	if diff == 0:
		diff = 1

	params = {
		"from_date": from_date,
		"to_date": to_date,
		"prev_from_date": frappe.utils.add_days(from_date, -diff),
	}

	if users:
		conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			COUNT(CASE
				WHEN d.creation >= %(from_date)s AND d.creation < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
					AND s.type NOT IN ('Won', 'Lost')
					{conds}
				THEN d.name
				ELSE NULL
			END) as current_month_deals,

			COUNT(CASE
				WHEN d.creation >= %(prev_from_date)s AND d.creation < %(from_date)s
					AND s.type NOT IN ('Won', 'Lost')
					{conds}
				THEN d.name
				ELSE NULL
			END) as prev_month_deals
		FROM `tabCRM Deal` d
		JOIN `tabCRM Deal Status` s ON d.status = s.name
	""",
		params,
		as_dict=1,
	)

	current_month_deals = result[0].current_month_deals or 0
	prev_month_deals = result[0].prev_month_deals or 0

	delta_in_percentage = (
		(current_month_deals - prev_month_deals) / prev_month_deals * 100 if prev_month_deals else 0
	)

	return {
		"title": _("Ongoing deals"),
		"tooltip": _("Total number of non won/lost deals"),
		"value": current_month_deals,
		"delta": delta_in_percentage,
		"deltaSuffix": "%",
	}


def get_average_ongoing_deal_value(from_date, to_date, users=None):
	"""
	Get ongoing deal count for the dashboard, and also calculate average deal value for ongoing deals.
	"""
	conds = ""

	diff = frappe.utils.date_diff(to_date, from_date)
	if diff == 0:
		diff = 1

	params = {
		"from_date": from_date,
		"to_date": to_date,
		"prev_from_date": frappe.utils.add_days(from_date, -diff),
	}

	if users:
		conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			AVG(CASE
				WHEN d.creation >= %(from_date)s AND d.creation < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
					AND s.type NOT IN ('Won', 'Lost')
					{conds}
				THEN d.deal_value * IFNULL(d.exchange_rate, 1)
				ELSE NULL
			END) as current_month_avg_value,

			AVG(CASE
				WHEN d.creation >= %(prev_from_date)s AND d.creation < %(from_date)s
					AND s.type NOT IN ('Won', 'Lost')
					{conds}
				THEN d.deal_value * IFNULL(d.exchange_rate, 1)
				ELSE NULL
			END) as prev_month_avg_value
		FROM `tabCRM Deal` d
		JOIN `tabCRM Deal Status` s ON d.status = s.name
    """,
		params,
		as_dict=1,
	)

	current_month_avg_value = result[0].current_month_avg_value or 0
	prev_month_avg_value = result[0].prev_month_avg_value or 0

	avg_value_delta = current_month_avg_value - prev_month_avg_value if prev_month_avg_value else 0

	return {
		"title": _("Avg. ongoing deal value"),
		"tooltip": _("Average deal value of non won/lost deals"),
		"value": current_month_avg_value,
		"delta": avg_value_delta,
		"prefix": get_base_currency_symbol(),
	}


def get_won_deals(from_date, to_date, users=None):
	"""
	Get won deal count for the dashboard, and also calculate average deal value for won deals.
	"""

	diff = frappe.utils.date_diff(to_date, from_date)
	if diff == 0:
		diff = 1

	conds = ""
	params = {
		"from_date": from_date,
		"to_date": to_date,
		"prev_from_date": frappe.utils.add_days(from_date, -diff),
	}

	if users:
		conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			COUNT(CASE
				WHEN d.closed_date >= %(from_date)s AND d.closed_date < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
					AND s.type = 'Won'
					{conds}
				THEN d.name
				ELSE NULL
			END) as current_month_deals,

			COUNT(CASE
				WHEN d.closed_date >= %(prev_from_date)s AND d.closed_date < %(from_date)s
					AND s.type = 'Won'
					{conds}
				THEN d.name
				ELSE NULL
			END) as prev_month_deals
		FROM `tabCRM Deal` d
		JOIN `tabCRM Deal Status` s ON d.status = s.name
		""",
		params,
		as_dict=1,
	)

	current_month_deals = result[0].current_month_deals or 0
	prev_month_deals = result[0].prev_month_deals or 0

	delta_in_percentage = (
		(current_month_deals - prev_month_deals) / prev_month_deals * 100 if prev_month_deals else 0
	)

	return {
		"title": _("Won deals"),
		"tooltip": _("Total number of won deals based on its closure date"),
		"value": current_month_deals,
		"delta": delta_in_percentage,
		"deltaSuffix": "%",
	}


def get_average_won_deal_value(from_date, to_date, users=None):
	"""
	Get won deal count for the dashboard, and also calculate average deal value for won deals.
	"""

	diff = frappe.utils.date_diff(to_date, from_date)
	if diff == 0:
		diff = 1

	conds = ""
	params = {
		"from_date": from_date,
		"to_date": to_date,
		"prev_from_date": frappe.utils.add_days(from_date, -diff),
	}

	if users:
		conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			AVG(CASE
				WHEN d.closed_date >= %(from_date)s AND d.closed_date < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
					AND s.type = 'Won'
					{conds}
				THEN d.deal_value * IFNULL(d.exchange_rate, 1)
				ELSE NULL
			END) as current_month_avg_value,

			AVG(CASE
				WHEN d.closed_date >= %(prev_from_date)s AND d.closed_date < %(from_date)s
					AND s.type = 'Won'
					{conds}
				THEN d.deal_value * IFNULL(d.exchange_rate, 1)
				ELSE NULL
			END) as prev_month_avg_value
		FROM `tabCRM Deal` d
		JOIN `tabCRM Deal Status` s ON d.status = s.name
		""",
		params,
		as_dict=1,
	)

	current_month_avg_value = result[0].current_month_avg_value or 0
	prev_month_avg_value = result[0].prev_month_avg_value or 0

	avg_value_delta = current_month_avg_value - prev_month_avg_value if prev_month_avg_value else 0

	return {
		"title": _("Avg. won deal value"),
		"tooltip": _("Average deal value of won deals"),
		"value": current_month_avg_value,
		"delta": avg_value_delta,
		"prefix": get_base_currency_symbol(),
	}


def get_average_deal_value(from_date, to_date, users=None):
	"""
	Get average deal value for the dashboard.
	"""

	diff = frappe.utils.date_diff(to_date, from_date)
	if diff == 0:
		diff = 1

	conds = ""
	params = {
		"from_date": from_date,
		"to_date": to_date,
		"prev_from_date": frappe.utils.add_days(from_date, -diff),
	}

	if users:
		conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			AVG(CASE
				WHEN d.creation >= %(from_date)s AND d.creation < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
					AND s.type != 'Lost'
					{conds}
				THEN d.deal_value * IFNULL(d.exchange_rate, 1)
				ELSE NULL
			END) as current_month_avg,

			AVG(CASE
				WHEN d.creation >= %(prev_from_date)s AND d.creation < %(from_date)s
					AND s.type != 'Lost'
					{conds}
				THEN d.deal_value * IFNULL(d.exchange_rate, 1)
				ELSE NULL
			END) as prev_month_avg
		FROM `tabCRM Deal` AS d
		JOIN `tabCRM Deal Status` s ON d.status = s.name
		""",
		params,
		as_dict=1,
	)

	current_month_avg = result[0].current_month_avg or 0
	prev_month_avg = result[0].prev_month_avg or 0

	delta = current_month_avg - prev_month_avg if prev_month_avg else 0

	return {
		"title": _("Avg. deal value"),
		"tooltip": _("Average deal value of ongoing & won deals"),
		"value": current_month_avg,
		"prefix": get_base_currency_symbol(),
		"delta": delta,
		"deltaSuffix": "%",
	}


def get_average_time_to_close_a_lead(from_date, to_date, users=None):
	"""
	Get average time to close deals for the dashboard.
	"""

	diff = frappe.utils.date_diff(to_date, from_date)
	if diff == 0:
		diff = 1

	conds = ""
	params = {
		"from_date": from_date,
		"to_date": to_date,
		"prev_from_date": frappe.utils.add_days(from_date, -diff),
		"prev_to_date": from_date,
	}

	if users:
		conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			AVG(CASE WHEN d.closed_date >= %(from_date)s AND d.closed_date < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
				THEN TIMESTAMPDIFF(DAY, COALESCE(l.creation, d.creation), d.closed_date) END) as current_avg_lead,
			AVG(CASE WHEN d.closed_date >= %(prev_from_date)s AND d.closed_date < %(prev_to_date)s
				THEN TIMESTAMPDIFF(DAY, COALESCE(l.creation, d.creation), d.closed_date) END) as prev_avg_lead
		FROM `tabCRM Deal` AS d
		JOIN `tabCRM Deal Status` s ON d.status = s.name
		LEFT JOIN `tabCRM Lead` l ON d.lead = l.name
		WHERE d.closed_date IS NOT NULL AND s.type = 'Won'
			{conds}
		""",
		params,
		as_dict=1,
	)

	current_avg_lead = result[0].current_avg_lead or 0
	prev_avg_lead = result[0].prev_avg_lead or 0
	delta_lead = current_avg_lead - prev_avg_lead if prev_avg_lead else 0

	return {
		"title": _("Avg. time to close a lead"),
		"tooltip": _("Average time taken from lead creation to deal closure"),
		"value": current_avg_lead,
		"suffix": " days",
		"delta": delta_lead,
		"deltaSuffix": " days",
		"negativeIsBetter": True,
	}


def get_average_time_to_close_a_deal(from_date, to_date, users=None):
	"""
	Get average time to close deals for the dashboard.
	"""

	diff = frappe.utils.date_diff(to_date, from_date)
	if diff == 0:
		diff = 1

	conds = ""
	params = {
		"from_date": from_date,
		"to_date": to_date,
		"prev_from_date": frappe.utils.add_days(from_date, -diff),
		"prev_to_date": from_date,
	}

	if users:
		conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			AVG(CASE WHEN d.closed_date >= %(from_date)s AND d.closed_date < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
				THEN TIMESTAMPDIFF(DAY, d.creation, d.closed_date) END) as current_avg_deal,
			AVG(CASE WHEN d.closed_date >= %(prev_from_date)s AND d.closed_date < %(prev_to_date)s
				THEN TIMESTAMPDIFF(DAY, d.creation, d.closed_date) END) as prev_avg_deal
		FROM `tabCRM Deal` AS d
		JOIN `tabCRM Deal Status` s ON d.status = s.name
		LEFT JOIN `tabCRM Lead` l ON d.lead = l.name
		WHERE d.closed_date IS NOT NULL AND s.type = 'Won'
			{conds}
		""",
		params,
		as_dict=1,
	)

	current_avg_deal = result[0].current_avg_deal or 0
	prev_avg_deal = result[0].prev_avg_deal or 0
	delta_deal = current_avg_deal - prev_avg_deal if prev_avg_deal else 0

	return {
		"title": _("Avg. time to close a deal"),
		"tooltip": _("Average time taken from deal creation to deal closure"),
		"value": current_avg_deal,
		"suffix": " days",
		"delta": delta_deal,
		"deltaSuffix": " days",
		"negativeIsBetter": True,
	}


def get_follow_up_leads(from_date, to_date, users=None):
    conds = ""
    params = {
        "from_date": from_date,
        "to_date": to_date,
    }

    if users:
        conds += " AND lead_owner IN %(users)s"
        params["users"] = tuple(users)

    result = frappe.db.sql(
        f"""
        SELECT COUNT(name) AS follow_up_count
        FROM `tabCRM Lead`
        WHERE status = 'Follow Up'
          AND creation >= %(from_date)s
          AND creation < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
          {conds}
        """,
        params,
        as_dict=True,
    )

    return {
        "title": _("Follow Up"),
        "tooltip": _("Total number of leads in Follow Up status"),
        "value": result[0].follow_up_count or 0,
    }


@frappe.whitelist()
# def get_call_lifecycle_sunburst(from_date, to_date, user=""):
#     conds = ""
#     params = {"from_date": from_date, "to_date": to_date}
#     if user:
#         conds += " AND owner = %(user)s"
#         params["user"] = user

#     # Groups by type and status for a hierarchical view
#     result = frappe.db.sql(f"""
#         SELECT type, status, COUNT(name) as count
#         FROM `tabCRM Call Log`
#         WHERE creation >= %(from_date)s AND creation < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
#         {conds}
#         GROUP BY type, status
#     """, params, as_dict=True)

#     return {
#         "title": _("Call Status Breakdown"),
#         "subtitle": _("Call distribution by Status"),
#         "data": result,
#         "type": "sunburst"
#     }

# updated function for percentage cal

def get_call_lifecycle_sunburst(from_date, to_date, users=None):
    conds = ""
    params = {"from_date": from_date, "to_date": to_date}

    if users:
        conds += " AND (caller IN %(users)s OR receiver IN %(users)s)"
        params["users"] = tuple(users)

    # Total calls
    total_calls = frappe.db.sql(f"""
        SELECT COUNT(name) as total
        FROM `tabCRM Call Log`
        WHERE creation >= %(from_date)s 
        AND creation < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
        {conds}
    """, params, as_dict=True)[0].total or 1

    result = frappe.db.sql(f"""
        SELECT status, COUNT(name) as count
        FROM `tabCRM Call Log`
        WHERE creation >= %(from_date)s 
        AND creation < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
        {conds}
        GROUP BY status
    """, params, as_dict=True)

    # percentage
    for row in result:
        row["percentage"] = round((row["count"] / total_calls) * 100, 2)

    return {
        "title": _("Call Status Breakdown"),
        "subtitle": _("Call distribution by Status"),
        "data": result,
        "type": "pie",  
        "total_calls": total_calls
    }


	# ------------------------------

@frappe.whitelist()
def get_call_volume_data(from_date, to_date, users=None):
    """
    Returns total counts for the bar chart to ensure parity with Call Insights.
    """
    conds = ""
    params = {"from_date": from_date, "to_date": to_date}
    if users:
        conds += " AND (caller IN %(users)s OR receiver IN %(users)s)"
        params["users"] = tuple(users)

    # Fetch counts grouped by type to get accurate totals
    result = frappe.db.sql(f"""
        SELECT 
            type, 
            COUNT(name) as count
        FROM `tabCRM Call Log`
        WHERE creation >= %(from_date)s AND creation < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
        {conds}
        GROUP BY type
    """, params, as_dict=True)

    # Transform into a simple dictionary for the frontend
    counts = { "Incoming": 0, "Outgoing": 0, "Total": 0 }
    for row in result:
        if row.type in counts:
            counts[row.type] = row.count
    
    counts["Total"] = counts["Incoming"] + counts["Outgoing"]

    return {
        "title": _("Call Breakdown"),
        "subtitle": _("Total incoming vs outgoing calls"),
        "data": counts,
        "type": "bar_volume"
    }


def get_sales_trend(from_date="", to_date="", users=None):
	"""
	Get sales trend data for the dashboard.
	[
		{ date: new Date('2024-05-01'), leads: 45, deals: 23, won_deals: 12 },
		{ date: new Date('2024-05-02'), leads: 50, deals: 30, won_deals: 15 },
		...
	]
	"""

	lead_conds = ""
	deal_conds = ""

	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	params = {"from": from_date, "to": to_date}

	if users:
		lead_conds += " AND lead_owner IN %(users)s"
		deal_conds += " AND deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			DATE_FORMAT(date, '%%Y-%%m-%%d') AS date,
			SUM(leads) AS leads,
			SUM(deals) AS deals,
			SUM(won_deals) AS won_deals
		FROM (
			SELECT
				DATE(creation) AS date,
				COUNT(*) AS leads,
				0 AS deals,
				0 AS won_deals
			FROM `tabCRM Lead`
			WHERE DATE(creation) BETWEEN %(from)s AND %(to)s
			{lead_conds}
			GROUP BY DATE(creation)

			UNION ALL

			SELECT
				DATE(d.creation) AS date,
				0 AS leads,
				COUNT(*) AS deals,
				SUM(CASE WHEN s.type = 'Won' THEN 1 ELSE 0 END) AS won_deals
			FROM `tabCRM Deal` d
			JOIN `tabCRM Deal Status` s ON d.status = s.name
			WHERE DATE(d.creation) BETWEEN %(from)s AND %(to)s
			{deal_conds}
			GROUP BY DATE(d.creation)
		) AS daily
		GROUP BY date
		ORDER BY date
		""",
		params,
		as_dict=True,
	)

	sales_trend = [
		{
			"date": frappe.utils.get_datetime(row.date).strftime("%Y-%m-%d"),
			"leads": row.leads or 0,
			"deals": row.deals or 0,
			"won_deals": row.won_deals or 0,
		}
		for row in result
	]

	return {
		"data": sales_trend,
		"title": _("Sales trend"),
		"subtitle": _("Daily performance of leads, deals, and wins"),
		"xAxis": {
			"title": _("Date"),
			"key": "date",
			"type": "time",
			"timeGrain": "day",
		},
		"yAxis": {
			"title": _("Count"),
		},
		"series": [
			{"name": "leads", "type": "line", "showDataPoints": True},
			{"name": "deals", "type": "line", "showDataPoints": True},
			{"name": "won_deals", "type": "line", "showDataPoints": True},
		],
	}


def get_forecasted_revenue(from_date="", to_date="", users=None):
	"""
	Get forecasted revenue for the dashboard.
	[
		{ date: new Date('2024-05-01'), forecasted: 1200000, actual: 980000 },
		{ date: new Date('2024-06-01'), forecasted: 1350000, actual: 1120000 },
		{ date: new Date('2024-07-01'), forecasted: 1600000, actual: "" },
		{ date: new Date('2024-08-01'), forecasted: 1500000, actual: "" },
		...
	]
	"""
	deal_conds = ""
	params = {}

	if users:
		deal_conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			DATE_FORMAT(d.expected_closure_date, '%%Y-%%m')                        AS month,
			SUM(
				CASE
					WHEN s.type = 'Lost' THEN d.expected_deal_value * IFNULL(d.exchange_rate, 1)
					ELSE d.expected_deal_value * IFNULL(d.probability, 0) / 100 * IFNULL(d.exchange_rate, 1)  -- forecasted
				END
			)                                                       AS forecasted,
			SUM(
				CASE
					WHEN s.type = 'Won' THEN d.deal_value * IFNULL(d.exchange_rate, 1)            -- actual
					ELSE 0
				END
			)                                                       AS actual
		FROM `tabCRM Deal` AS d
		JOIN `tabCRM Deal Status` s ON d.status = s.name
		WHERE d.expected_closure_date >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
		{deal_conds}
		GROUP BY DATE_FORMAT(d.expected_closure_date, '%%Y-%%m')
		ORDER BY month
		""",
		params,
		as_dict=True,
	)

	for row in result:
		row["month"] = frappe.utils.get_datetime(row["month"]).strftime("%Y-%m-01")
		row["forecasted"] = row["forecasted"] or ""
		row["actual"] = row["actual"] or ""

	return {
		"data": result or [],
		"title": _("Forecasted revenue"),
		"subtitle": _("Projected vs actual revenue based on deal probability"),
		"xAxis": {
			"title": _("Month"),
			"key": "month",
			"type": "time",
			"timeGrain": "month",
		},
		"yAxis": {
			"title": _("Revenue") + f" ({get_base_currency_symbol()})",
		},
		"series": [
			{"name": "forecasted", "type": "line", "showDataPoints": True},
			{"name": "actual", "type": "line", "showDataPoints": True},
		],
	}


def get_funnel_conversion(from_date="", to_date="", users=None):
	"""
	Get funnel conversion data for the dashboard.
	[
		{ stage: 'Leads', count: 120 },
		{ stage: 'Qualification', count: 100 },
		{ stage: 'Negotiation', count: 80 },
		{ stage: 'Ready to Close', count: 60 },
		{ stage: 'Won', count: 30 },
		...
	]
	"""
	lead_conds = ""
	deal_conds = ""

	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	lead_filters = {"from": from_date, "to": to_date}
	deal_filters = {"from": from_date, "to": to_date}

	if users:
		lead_conds += " AND lead_owner IN %(users)s"
		deal_conds += " AND deal_owner IN %(users)s"
		lead_filters["users"] = tuple(users)
		deal_filters["users"] = tuple(users)

	result = []

	# Get total leads
	total_leads = frappe.db.sql(
		f"""
			SELECT COUNT(*) AS count
			FROM `tabCRM Lead`
			WHERE DATE(creation) BETWEEN %(from)s AND %(to)s
			{lead_conds}
		""",
		lead_filters,
		as_dict=True,
	)
	total_leads_count = total_leads[0].count if total_leads else 0

	result.append({"stage": "Leads", "count": total_leads_count})

	result += get_deal_status_change_counts(from_date, to_date, deal_conds, deal_filters)

	return {
		"data": result or [],
		"title": _("Funnel conversion"),
		"subtitle": _("Lead to deal conversion pipeline"),
		"xAxis": {
			"title": _("Stage"),
			"key": "stage",
			"type": "category",
		},
		"yAxis": {
			"title": _("Count"),
		},
		"swapXY": True,
		"series": [
			{
				"name": "count",
				"type": "bar",
				"echartOptions": {
					"colorBy": "data",
				},
			},
		],
	}


def get_deals_by_stage_axis(from_date="", to_date="", users=None):
	"""
	Get deal data by stage for the dashboard.
	[
		{ stage: 'Prospecting', count: 120 },
		{ stage: 'Negotiation', count: 45 },
		...
	]
	"""
	deal_conds = ""

	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	params = {"from": from_date, "to": to_date}

	if users:
		deal_conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			d.status AS stage,
			COUNT(*) AS count,
			s.type AS status_type
		FROM `tabCRM Deal` AS d
		JOIN `tabCRM Deal Status` s ON d.status = s.name
		WHERE DATE(d.creation) BETWEEN %(from)s AND %(to)s AND s.type NOT IN ('Lost')
		{deal_conds}
		GROUP BY d.status
		ORDER BY count DESC
		""",
		params,
		as_dict=True,
	)

	return {
		"data": result or [],
		"title": _("Deals by ongoing & won stage"),
		"xAxis": {
			"title": _("Stage"),
			"key": "stage",
			"type": "category",
		},
		"yAxis": {"title": _("Count")},
		"series": [
			{"name": "count", "type": "bar"},
		],
	}


def get_deals_by_stage_donut(from_date="", to_date="", users=None):
	"""
	Get deal data by stage for the dashboard.
	[
		{ stage: 'Prospecting', count: 120 },
		{ stage: 'Negotiation', count: 45 },
		...
	]
	"""
	deal_conds = ""

	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	params = {"from": from_date, "to": to_date}

	if users:
		deal_conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			d.status AS stage,
			COUNT(*) AS count,
			s.type AS status_type
		FROM `tabCRM Deal` AS d
		JOIN `tabCRM Deal Status` s ON d.status = s.name
		WHERE DATE(d.creation) BETWEEN %(from)s AND %(to)s
		{deal_conds}
		GROUP BY d.status
		ORDER BY count DESC
		""",
		params,
		as_dict=True,
	)

	return {
		"data": result or [],
		"title": _("Deals by stage"),
		"subtitle": _("Current pipeline distribution"),
		"categoryColumn": "stage",
		"valueColumn": "count",
	}


def get_lost_deal_reasons(from_date="", to_date="", users=None):
	"""
	Get lost deal reasons for the dashboard.
	[
		{ reason: 'Price too high', count: 20 },
		{ reason: 'Competitor won', count: 15 },
		...
	]
	"""

	deal_conds = ""

	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	params = {"from": from_date, "to": to_date}

	if users:
		deal_conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			d.lost_reason AS reason,
			COUNT(*) AS count
		FROM `tabCRM Deal` AS d
		JOIN `tabCRM Deal Status` s ON d.status = s.name
		WHERE DATE(d.creation) BETWEEN %(from)s AND %(to)s AND s.type = 'Lost'
		{deal_conds}
		GROUP BY d.lost_reason
		HAVING reason IS NOT NULL AND reason != ''
		ORDER BY count DESC
		""",
		params,
		as_dict=True,
	)

	return {
		"data": result or [],
		"title": _("Lost deal reasons"),
		"subtitle": _("Common reasons for losing deals"),
		"xAxis": {
			"title": _("Reason"),
			"key": "reason",
			"type": "category",
		},
		"yAxis": {
			"title": _("Count"),
		},
		"series": [
			{"name": "count", "type": "bar"},
		],
	}


def get_leads_by_source(from_date="", to_date="", users=None):
	"""
	Get lead data by source for the dashboard.
	[
		{ source: 'Website', count: 120 },
		{ source: 'Referral', count: 45 },
		...
	]
	"""
	lead_conds = ""

	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	params = {"from": from_date, "to": to_date}

	if users:
		lead_conds += " AND lead_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			IFNULL(source, 'Empty') AS source,
			COUNT(*) AS count
		FROM `tabCRM Lead`
		WHERE DATE(creation) BETWEEN %(from)s AND %(to)s
		{lead_conds}
		GROUP BY source
		ORDER BY count DESC
		""",
		params,
		as_dict=True,
	)

	return {
		"data": result or [],
		"title": _("Leads by source"),
		"subtitle": _("Lead generation channel analysis"),
		"categoryColumn": "source",
		"valueColumn": "count",
	}


def get_deals_by_source(from_date="", to_date="", users=None):
	"""
	Get deal data by source for the dashboard.
	[
		{ source: 'Website', count: 120 },
		{ source: 'Referral', count: 45 },
		...
	]
	"""
	deal_conds = ""

	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	params = {"from": from_date, "to": to_date}

	if users:
		deal_conds += " AND deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			IFNULL(source, 'Empty') AS source,
			COUNT(*) AS count
		FROM `tabCRM Deal`
		WHERE DATE(creation) BETWEEN %(from)s AND %(to)s
		{deal_conds}
		GROUP BY source
		ORDER BY count DESC
		""",
		params,
		as_dict=True,
	)

	return {
		"data": result or [],
		"title": _("Deals by source"),
		"subtitle": _("Deal generation channel analysis"),
		"categoryColumn": "source",
		"valueColumn": "count",
	}


def get_deals_by_territory(from_date="", to_date="", users=None):
	"""
	Get deal data by territory for the dashboard.
	[
		{ territory: 'North America', deals: 45, value: 2300000 },
		{ territory: 'Europe', deals: 30, value: 1500000 },
		...
	]
	"""
	deal_conds = ""

	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	params = {"from": from_date, "to": to_date}

	if users:
		deal_conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			IFNULL(d.territory, 'Empty') AS territory,
			COUNT(*) AS deals,
			SUM(COALESCE(d.deal_value, 0) * IFNULL(d.exchange_rate, 1)) AS value
		FROM `tabCRM Deal` AS d
		WHERE DATE(d.creation) BETWEEN %(from)s AND %(to)s
		{deal_conds}
		GROUP BY d.territory
		ORDER BY deals DESC, value DESC
		""",
		params,
		as_dict=True,
	)

	return {
		"data": result or [],
		"title": _("Deals by territory"),
		"subtitle": _("Geographic distribution of deals and revenue"),
		"xAxis": {
			"title": _("Territory"),
			"key": "territory",
			"type": "category",
		},
		"yAxis": {
			"title": _("Number of deals"),
		},
		"y2Axis": {
			"title": _("Deal value") + f" ({get_base_currency_symbol()})",
		},
		"series": [
			{"name": "deals", "type": "bar"},
			{"name": "value", "type": "line", "showDataPoints": True, "axis": "y2"},
		],
	}


def get_deals_by_salesperson(from_date="", to_date="", users=None):
	"""
	Get deal data by salesperson for the dashboard.
	[
		{ salesperson: 'John Smith', deals: 45, value: 2300000 },
		{ salesperson: 'Jane Doe', deals: 30, value: 1500000 },
		...
	]
	"""
	deal_conds = ""

	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	params = {"from": from_date, "to": to_date}

	if users:
		deal_conds += " AND d.deal_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
			IFNULL(u.full_name, d.deal_owner) AS salesperson,
			COUNT(*)                           AS deals,
			SUM(COALESCE(d.deal_value, 0) * IFNULL(d.exchange_rate, 1)) AS value
		FROM `tabCRM Deal` AS d
		LEFT JOIN `tabUser` AS u ON u.name = d.deal_owner
		WHERE DATE(d.creation) BETWEEN %(from)s AND %(to)s
		{deal_conds}
		GROUP BY d.deal_owner
		ORDER BY deals DESC, value DESC
		""",
		params,
		as_dict=True,
	)

	return {
		"data": result or [],
		"title": _("Deals by salesperson"),
		"subtitle": _("Number of deals and total value per salesperson"),
		"xAxis": {
			"title": _("Salesperson"),
			"key": "salesperson",
			"type": "category",
		},
		"yAxis": {
			"title": _("Number of deals"),
		},
		"y2Axis": {
			"title": _("Deal value") + f" ({get_base_currency_symbol()})",
		},
		"series": [
			{"name": "deals", "type": "bar"},
			{"name": "value", "type": "line", "showDataPoints": True, "axis": "y2"},
		],
	}


def get_base_currency_symbol():
	"""
	Get the base currency symbol from the system settings.
	"""
	base_currency = frappe.db.get_single_value("FCRM Settings", "currency") or "USD"
	return frappe.db.get_value("Currency", base_currency, "symbol") or ""


def get_deal_status_change_counts(from_date, to_date, deal_conds="", filters=None):
	"""
	Get count of each status change (to) for each deal, excluding deals with current status type 'Lost'.
	Order results by status position.
	Returns:
	[
	  {"status": "Qualification", "count": 120},
	  {"status": "Negotiation", "count": 85},
	  ...
	]
	"""
	params = (filters or {}).copy()
	params.setdefault("from", from_date)
	params.setdefault("to", to_date)

	result = frappe.db.sql(
		f"""
		SELECT
			scl.to AS stage,
			COUNT(*) AS count
		FROM
			`tabCRM Status Change Log` scl
		JOIN
			`tabCRM Deal` d ON scl.parent = d.name
		JOIN
			`tabCRM Deal Status` s ON d.status = s.name
		JOIN
			`tabCRM Deal Status` st ON scl.to = st.name
		WHERE
			scl.to IS NOT NULL
			AND scl.to != ''
			AND s.type != 'Lost'
			AND DATE(d.creation) BETWEEN %(from)s AND %(to)s
			{deal_conds}
		GROUP BY
			scl.to, st.position
		ORDER BY
			st.position ASC
		""",
		params,
		as_dict=True,
	)
	return result or []


# ===== CUSTOM DASHBOARD FUNCTIONS =====

def get_fresh_leads(from_date, to_date, users=None):
	"""
	Get fresh leads count (leads created today) for the dashboard.
	"""
	conds = ""
	today = frappe.utils.nowdate()
	yesterday = frappe.utils.add_days(today, -1)
	
	params = {
		"today": today,
		"yesterday": yesterday,
	}

	if users:
		conds += " AND lead_owner IN %(users)s"
		params["users"] = tuple(users)

	result = frappe.db.sql(
		f"""
		SELECT
            COUNT(CASE
                WHEN DATE(creation) = %(today)s
                {conds}
                THEN name
                ELSE NULL
            END) as today_leads,

            COUNT(CASE
                WHEN DATE(creation) = %(yesterday)s
                {conds}
                THEN name
                ELSE NULL
            END) as yesterday_leads
		FROM `tabCRM Lead`
	""",
		params,
		as_dict=1,
	)

	today_leads = result[0].today_leads or 0
	yesterday_leads = result[0].yesterday_leads or 0

	delta = today_leads - yesterday_leads

	return {
		"title": _("Fresh leads"),
		"tooltip": _("Leads created today"),
		"value": today_leads,
		"delta": delta,
		"deltaSuffix": "",
	}


@frappe.whitelist()
def get_call_insights(from_date, to_date, users=None):
    conds = ""
    params = {
        "from_date": from_date,
        "to_date": to_date,
    }

    if users:
        conds += " AND (caller IN %(users)s OR receiver IN %(users)s)"
        params["users"] = tuple(users)

    result = frappe.db.sql(
        f"""
        SELECT
            COUNT(name) AS total_calls,

            SUM(type = 'Incoming') AS incoming_calls,
            SUM(type = 'Outgoing') AS outgoing_calls,

            SUM(status = 'Ringing') AS ringing_calls,
            SUM(status = 'Completed') AS completed_calls,
            SUM(status = 'Failed') AS failed_calls,
            SUM(status = 'Busy') AS busy_calls,
            SUM(status = 'Queued') AS queued_calls,
            SUM(status IN ('Canceled', 'Cancelled')) AS canceled_calls,
            SUM(status = 'Call not receive by agent') AS not_received_agent_calls,
            SUM(status = 'Call not receive by agent (Over Smartphone)') AS not_received_agent_smartphone_calls,
            SUM(status = 'Not received by seller') AS not_received_seller_calls,

            IFNULL(SUM(duration), 0) AS total_duration

        FROM `tabCRM Call Log`
        WHERE creation >= %(from_date)s
          AND creation < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
          {conds}
        """,
        params,
        as_dict=True,
    )[0]

    def format_duration(seconds):
        seconds = int(seconds or 0)
        minutes = seconds // 60
        secs = seconds % 60
        return f"{minutes} min {secs} sec"

    def card(label, value):
        """Only include card if value > 0"""
        return {"label": _(label), "value": value or 0, "hidden": not bool(value)}

    return {
        "title": _("Call Insights"),
        "data": [
            {"label": _("Total Calls"), "value": result.total_calls or 0, "hidden": False},
            {"label": _("Incoming Calls"), "value": result.incoming_calls or 0, "hidden": False},
            {"label": _("Outgoing Calls"), "value": result.outgoing_calls or 0, "hidden": False},

            card("Ringing", result.ringing_calls),
            card("Completed", result.completed_calls),
            card("Failed", result.failed_calls),
            card("Busy", result.busy_calls),
            card("Queued", result.queued_calls),
            card("Canceled", result.canceled_calls),
            card("Not Received by Agent", result.not_received_agent_calls),
            card("Not Received (Smartphone)", result.not_received_agent_smartphone_calls),
            card("Not Received by Seller", result.not_received_seller_calls),

            {
                "label": _("Total Talk Time"),
                "value": format_duration(result.total_duration),
                "hidden": False,
            },
        ],
    }


def get_followup_insights(from_date="", to_date="", users=None):
	"""
	Get follow-up insights for the dashboard.
	Returns counts for:
	- Planned: Future follow-ups
	- Pending: Due today or overdue but not completed
	- Rescheduled: Follow-ups that were rescheduled
	- Cancelled: Follow-ups that were cancelled
	- Done: Completed follow-ups
	- Missed: Overdue and not completed
	"""
	
	lead_conds = ""
	params = {
		"today": frappe.utils.nowdate(),
	}
	
	if users:
		lead_conds += " AND lead_owner IN %(users)s"
		params["users"] = tuple(users)
	
	# Get counts for each follow-up status (ALL follow-ups, not filtered by date range)
	# Only count unconverted leads to match the Leads list view
	result = frappe.db.sql(
		f"""
		SELECT
			COUNT(CASE 
				WHEN followup_status = 'Planned'
				THEN 1 
			END) as planned,
			
			COUNT(CASE 
				WHEN followup_status = 'Pending'
				THEN 1 
			END) as pending,
			
			COUNT(CASE 
				WHEN followup_status = 'Rescheduled' 
				THEN 1 
			END) as rescheduled,
			
			COUNT(CASE 
				WHEN followup_status = 'Cancelled' 
				THEN 1 
			END) as cancelled,
			
			COUNT(CASE 
				WHEN followup_status = 'Done' 
				THEN 1 
			END) as done,
			
			COUNT(CASE 
				WHEN followup_status = 'Missed'
				THEN 1 
			END) as missed,
			
			COUNT(CASE 
				WHEN next_followup_date IS NOT NULL 
				THEN 1 
			END) as total
		FROM `tabCRM Lead`
		WHERE next_followup_date IS NOT NULL
		AND converted = 0
		{lead_conds}
		""",
		params,
		as_dict=True,
	)
	
	data = result[0] if result else {
		"planned": 0,
		"pending": 0,
		"rescheduled": 0,
		"cancelled": 0,
		"done": 0,
		"missed": 0,
		"total": 0,
	}
	
	return {
		"title": _("Follow-Up Insights"),
		"data": [
			{
				"label": _("Planned"),
				"value": data.get("planned", 0),
				"status": "Planned",
				"icon": "calendar-check",
				"color": "blue",
			},
			{
				"label": _("Pending"),
				"value": data.get("pending", 0),
				"status": "Pending",
				"icon": "clock",
				"color": "orange",
			},
			{
				"label": _("Rescheduled"),
				"value": data.get("rescheduled", 0),
				"status": "Rescheduled",
				"icon": "calendar-clock",
				"color": "purple",
			},
			{
				"label": _("Cancelled"),
				"value": data.get("cancelled", 0),
				"status": "Cancelled",
				"icon": "x-circle",
				"color": "gray",
			},
			{
				"label": _("Done"),
				"value": data.get("done", 0),
				"status": "Done",
				"icon": "check-circle",
				"color": "green",
			},
			{
				"label": _("Missed"),
				"value": data.get("missed", 0),
				"status": "Missed",
				"icon": "alert-circle",
				"color": "red",
			},
		],
		"total": data.get("total", 0),
	}


	# ---------------------


def get_lead_action_insights(from_date, to_date, users=None):
	params = {
		"from_date": from_date,
		"to_date": f"{to_date} 23:59:59",
	}

	user_filter = ""
	if users:
		user_filter = " AND lead_owner IN %(users)s"
		params["users"] = tuple(users)

	total_assigned = (
		frappe.db.sql(
			f"""
		SELECT COUNT(name)
		FROM `tabCRM Lead`
		WHERE creation BETWEEN %(from_date)s AND %(to_date)s
		{user_filter}
	""",
			params,
		)[0][0]
		or 0
	)

	log_user_filter = ""
	if users:
		log_user_filter = " AND assigned_user IN %(users)s"

	# Check if tabLead Department Log exists before querying
	if not frappe.db.table_exists("Lead Department Log"):
		return {
			"data": [{"action": "Pending/No Action", "count": total_assigned}],
			"type": "lead_actions_donut",
			"title": _("Lead Action Breakdown"),
			"subtitle": _("Total Leads: {0}").format(total_assigned),
			"colors": ["#8B5CF6", "#22D3EE", "#34D399", "#FBBF24", "#F87171"],
		}

	action_data = frappe.db.sql(
		f"""
		SELECT
			CASE
				WHEN COALESCE(last_action, action) = 'Forward' THEN 'Mark Done'
				WHEN COALESCE(last_action, action) = 'Backward' THEN 'Send Back'
				WHEN COALESCE(last_action, action) = 'Reject' THEN 'Reject to Onboarding'
				WHEN COALESCE(last_action, action) = 'Manager Override' THEN 'Transfer to Department'
				ELSE COALESCE(last_action, action)
			END as label_action,
			COUNT(DISTINCT parent) as count
		FROM `tabLead Department Log`
		WHERE COALESCE(last_action, action) IS NOT NULL
		AND COALESCE(last_action, action) != 'Initial'
		AND creation BETWEEN %(from_date)s AND %(to_date)s
		{log_user_filter}
		GROUP BY label_action
	""",
		params,
		as_dict=True,
	)

	total_acted_upon = sum(d["count"] for d in action_data)
	no_action_count = max(0, total_assigned - total_acted_upon)

	chart_data = [{"action": d.label_action, "count": d.count} for d in action_data]
	if no_action_count > 0:
		chart_data.append({"action": "Pending/No Action", "count": no_action_count})
		
	mark_done_percent = 0
	mark_done_count = sum(d["count"] for d in action_data if d.label_action == "Mark Done")
	
	if total_assigned > 0:
		mark_done_percent = round((mark_done_count / total_assigned) * 100, 1)

	return {
		"data": chart_data,
		"type": "lead_actions_donut",
		"title": _("Lead Action Breakdown"),
		"subtitle": _("Total Leads: {0}").format(total_assigned),
		"mark_done_percent": mark_done_percent,
		"colors": ["#8B5CF6", "#22D3EE", "#34D399", "#FBBF24", "#F87171"],
	}


def get_lead_status_analytics(from_date, to_date, users=None):
	params = {"from_date": from_date, "to_date": f"{to_date} 23:59:59"}

	conditions = ["creation BETWEEN %(from_date)s AND %(to_date)s"]

	if users:
		conditions.append("lead_owner IN %(users)s")
		params["users"] = tuple(users)

	where_clause = " WHERE " + " AND ".join(conditions)

	data = frappe.db.sql(
		f"""
		SELECT status, COUNT(*) as count
		FROM `tabCRM Lead`
		{where_clause}
		GROUP BY status
		ORDER BY count DESC
	""",
		params,
		as_dict=True,
	)

	return {
		"type": "lead_status_distribution",
		"title": _("Lead Status Distribution"),
		"data": data,
		"colors": ["#FFDEA6", "#D3A4FF"],
	}


@frappe.whitelist()
def export_dashboard_data(from_date="", to_date="", user="", department=""):
	"""Export dashboard data as a formatted Excel file with per-user sheets."""
	import io
	import base64
	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill, Border, Side

	if not from_date or not to_date:
		from_date = frappe.utils.get_first_day(from_date or frappe.utils.nowdate())
		to_date = frappe.utils.get_last_day(to_date or frappe.utils.nowdate())

	roles = frappe.get_roles(frappe.session.user)
	is_sales_manager = "Sales Manager" in roles or "System Manager" in roles
	is_sales_user = "Sales User" in roles and not is_sales_manager

	all_users = []
	if is_sales_user:
		all_users = [frappe.session.user]
	elif department and not user:
		all_users = get_department_users(department)
	elif user:
		all_users = [user]

	# Styles
	section_font = Font(bold=True, color="FFFFFF", size=11)
	section_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
	bold_font = Font(bold=True, size=10)
	normal_font = Font(size=10)
	title_font = Font(bold=True, size=12)
	thin_border = Border(
		left=Side(style="thin", color="D9D9D9"),
		right=Side(style="thin", color="D9D9D9"),
		top=Side(style="thin", color="D9D9D9"),
		bottom=Side(style="thin", color="D9D9D9"),
	)

	wb = Workbook()

	def _write_sheet(ws, sheet_label, target_users):
		"""Write dashboard data for a set of users into a worksheet."""
		users_q = target_users if target_users else []
		if department and not target_users:
			users_q = ["__no_match__"]

		row = 1

		# Header
		ws.cell(row=row, column=1, value="Dashboard Export").font = title_font
		row += 1
		ws.cell(row=row, column=1, value=f"User: {sheet_label}").font = normal_font
		row += 1
		ws.cell(row=row, column=1, value=f"Period: {from_date} to {to_date}").font = normal_font
		row += 2

		# Dashboard Metrics section header
		for col in range(1, 4):
			c = ws.cell(row=row, column=col)
			c.fill = section_fill
			c.font = section_font
		ws.cell(row=row, column=1, value="Dashboard Metrics")
		row += 1

		for col, lbl in enumerate(["Metric", "Value", "Delta"], 1):
			c = ws.cell(row=row, column=col, value=lbl)
			c.font = bold_font
			c.border = thin_border
		row += 1

		metrics = [
			("Total Leads", "get_total_leads"),
			("Ongoing Deals", "get_ongoing_deals"),
			("Won Deals", "get_won_deals"),
			("Avg. Deal Value", "get_average_deal_value"),
			("Avg. Time to Close Lead", "get_average_time_to_close_a_lead"),
			("Avg. Time to Close Deal", "get_average_time_to_close_a_deal"),
			("Follow Up Leads", "get_follow_up_leads"),
			("Fresh Leads", "get_fresh_leads"),
		]

		for lbl, fn in metrics:
			try:
				mod = frappe.get_attr("crm.api.dashboard")
				if hasattr(mod, fn):
					d = getattr(mod, fn)(from_date, to_date, users_q)
					val = d.get("value", 0)
					dlt = d.get("delta", "")
					sfx = d.get("suffix", "")
					dsfx = d.get("deltaSuffix", "")
					pfx = d.get("prefix", "")
					v_str = f"{pfx}{val}{sfx}" if pfx or sfx else str(val)
					d_str = f"{dlt}{dsfx}" if dlt != "" else ""
				else:
					v_str, d_str = "N/A", ""
			except Exception:
				v_str, d_str = "Error", ""

			for col, v in enumerate([lbl, v_str, d_str], 1):
				c = ws.cell(row=row, column=col, value=v)
				c.font = normal_font
				c.border = thin_border
			row += 1

		row += 1

		# Call Insights section header
		for col in range(1, 3):
			c = ws.cell(row=row, column=col)
			c.fill = section_fill
			c.font = section_font
		ws.cell(row=row, column=1, value="Call Insights")
		row += 1

		for col, lbl in enumerate(["Metric", "Value"], 1):
			c = ws.cell(row=row, column=col, value=lbl)
			c.font = bold_font
			c.border = thin_border
		row += 1

		try:
			cd = get_call_insights(from_date, to_date, users_q)
			if cd and cd.get("data"):
				for item in cd["data"]:
					for col, v in enumerate([item.get("label", ""), item.get("value", 0)], 1):
						c = ws.cell(row=row, column=col, value=v)
						c.font = normal_font
						c.border = thin_border
					row += 1
		except Exception:
			ws.cell(row=row, column=1, value="Error").font = normal_font
			row += 1

		row += 1

		# Lead Data section header
		lead_headers = ["Lead Name", "Status", "Organization", "Email", "Mobile", "Created On"]
		for col in range(1, len(lead_headers) + 1):
			c = ws.cell(row=row, column=col)
			c.fill = section_fill
			c.font = section_font
		ws.cell(row=row, column=1, value="Lead Data")
		row += 1

		for col, lbl in enumerate(lead_headers, 1):
			c = ws.cell(row=row, column=col, value=lbl)
			c.font = bold_font
			c.border = thin_border
		row += 1

		lc = ""
		lp = {"from_date": from_date, "to_date": to_date}
		if users_q:
			lc += " AND lead_owner IN %(users)s"
			lp["users"] = tuple(users_q)

		leads = frappe.db.sql(
			f"""
			SELECT name, status, organization, email, mobile_no, creation
			FROM `tabCRM Lead`
			WHERE creation >= %(from_date)s AND creation < DATE_ADD(%(to_date)s, INTERVAL 1 DAY)
			{lc}
			ORDER BY creation DESC
			""",
			lp,
			as_dict=True,
		)

		for lead in leads:
			vals = [
				lead.get("name", ""),
				lead.get("status", ""),
				lead.get("organization", "") or "",
				lead.get("email", "") or "",
				lead.get("mobile_no", "") or "",
				str(lead.get("creation", "")),
			]
			for col, v in enumerate(vals, 1):
				c = ws.cell(row=row, column=col, value=v)
				c.font = normal_font
				c.border = thin_border
			row += 1

		# Auto-fit column widths
		for col_cells in ws.columns:
			max_len = 0
			col_letter = col_cells[0].column_letter
			for cell in col_cells:
				if cell.value:
					max_len = max(max_len, len(str(cell.value)))
			ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

	def _safe_sheet_name(name):
		"""Remove characters invalid in Excel sheet names."""
		import re
		# Excel forbids: \ / * ? : [ ]
		clean = re.sub(r'[\\/*?:\[\]]', '', name)
		return clean[:31] or "Sheet"

	# --- Build Workbook ---
	if department and not user:
		main_label = f"Department - {department}"
	elif user:
		main_label = frappe.db.get_value("User", user, "full_name") or user
	else:
		main_label = "All Users"

	# Main overview sheet
	ws_main = wb.active
	ws_main.title = _safe_sheet_name(main_label)
	_write_sheet(ws_main, main_label, all_users if all_users else [])

	# Per-user sheets when a department is selected
	if department and not user and all_users:
		for dept_user in all_users:
			full_name = frappe.db.get_value("User", dept_user, "full_name") or dept_user
			ws_user = wb.create_sheet(title=_safe_sheet_name(full_name))
			_write_sheet(ws_user, full_name, [dept_user])

	# Save to bytes and return as base64
	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)
	file_data = base64.b64encode(buf.getvalue()).decode("utf-8")
	buf.close()

	safe_filename = main_label.replace(' ', '_').replace(':', '-')
	return {
		"file_data": file_data,
		"file_name": f"CRM_Dashboard_{safe_filename}_{from_date}_to_{to_date}.xlsx",
	}

